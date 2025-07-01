"""
NCSS-Style Export Utilities

This module provides functions to export NCSS-style report data
to various machine-readable formats (CSV, Excel, JSON) for
regulatory compliance and data sharing purposes.

Functions:
    export_table_to_csv: Export single table to CSV
    export_table_to_excel: Export single table to Excel
    export_report_to_excel: Export entire report to Excel with multiple sheets
    export_report_to_json: Export report to JSON format
    export_tables_to_csv_batch: Export all tables to separate CSV files
    create_data_dictionary: Create data dictionary for all variables
    export_data_dictionary: Export data dictionary to CSV
    create_regulatory_summary: Create regulatory-compliant summary
    export_regulatory_summary: Export regulatory summary to JSON

Example:
    # Export report to Excel
    export_report_to_excel(report, "analysis_results.xlsx")
    
    # Create data dictionary
    data_dict = create_data_dictionary(report)
    export_data_dictionary(report, "data_dictionary.csv")
"""

import json
import base64
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import tempfile

from .ncss_report_structures import NCSSReport, NCSSSection, NCSSTable, NCSSPlot, SectionType
from .ncss_report_builder import format_ncss_table_rows


def export_table_to_csv(table: NCSSTable, filepath: str) -> None:
    """
    Export a single NCSS table to CSV format.
    
    Args:
        table: NCSS table to export
        filepath: Output file path
        
    Raises:
        FileNotFoundError: If directory doesn't exist and can't be created
        PermissionError: If file can't be written
    """
    try:
        df = table.to_dataframe()
        df.to_csv(filepath, index=False)
        print(f"Table '{table.title}' exported to {filepath}")
    except Exception as e:
        print(f"Failed to export table '{table.title}': {e}")
        raise


def export_table_to_excel(table: NCSSTable, filepath: str, sheet_name: Optional[str] = None) -> None:
    """
    Export a single NCSS table to Excel format.
    
    Args:
        table: NCSS table to export
        filepath: Output file path
        sheet_name: Optional sheet name (defaults to table title)
        
    Raises:
        FileNotFoundError: If directory doesn't exist and can't be created
        PermissionError: If file can't be written
    """
    try:
        # Format rows before export
        formatted_rows = format_ncss_table_rows(table.rows, table.columns)
        df = pd.DataFrame(formatted_rows)
        sheet_name = sheet_name or table.title[:31]  # Excel sheet name limit
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Table '{table.title}' exported to {filepath}")
    except Exception as e:
        print(f"Failed to export table '{table.title}': {e}")
        raise


def export_report_to_excel(report: NCSSReport, filepath: str, extra_plots: Optional[list] = None) -> None:
    """
    Export entire NCSS report to Excel with multiple sheets.
    Adds a 'Plots' sheet (last) with all diagnostic plots as PNGs.
    The 'Summary' sheet is currently disabled due to formatting issues and is left for future implementation.
    Args:
        report: NCSS report to export
        filepath: Output file path
        extra_plots: Optional list of NCSSPlot to add to the Plots sheet
    Raises:
        FileNotFoundError: If directory doesn't exist and can't be created
        PermissionError: If file can't be written
    """
    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # --- Summary sheet disabled for now ---
            # summary_rows = []
            # for section in report.sections:
            #     for table in section.tables:
            #         # Add section and table headers
            #         summary_rows.append({"Section": section.title, "Table": table.title})
            #         # Add table data
            #         formatted_rows = format_ncss_table_rows(table.rows, table.columns)
            #         df = pd.DataFrame(formatted_rows)
            #         if not df.empty:
            #             # Add a blank row before each table except the first
            #             if len(summary_rows) > 1:
            #                 summary_rows.append({})
            #             # Add column headers as a row
            #             summary_rows.append({col: col for col in table.columns})
            #             # Add table rows
            #             summary_rows.extend(formatted_rows)
            # if summary_rows:
            #     summary_df = pd.DataFrame(summary_rows)
            #     summary_df.to_excel(writer, sheet_name='Summary', index=False)
            # --- End summary sheet code ---
            # 2. Per-section sheets (as before)
            for section in report.sections:
                section_name = section.title[:31]  # Excel sheet name limit
                for i, table in enumerate(section.tables):
                    sheet_name = f"{section_name}_{i+1}"[:31] if len(section.tables) > 1 else section_name
                    formatted_rows = format_ncss_table_rows(table.rows, table.columns)
                    df = pd.DataFrame(formatted_rows)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.book.save(filepath)  # Save so openpyxl can open
        # 3. Plots sheet (after writing tables)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.create_sheet('Plots')
        row = 1
        # Gather all plots from report
        all_plots = []
        for section in report.sections:
            for plot in section.plots:
                all_plots.append(plot)
        if extra_plots:
            all_plots.extend(extra_plots)
        for plot in all_plots:
            # Write plot title
            ws.cell(row=row, column=1, value=plot.title)
            row += 1
            # Save image to temp file and insert
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_img:
                tmp_img.write(plot.image_bytes)
                tmp_img.flush()
                img = XLImage(tmp_img.name)
                img.width = 480
                img.height = 320
                ws.add_image(img, f'A{row}')
            row += 18  # Space between plots
        wb.save(filepath)
        print(f"Report '{report.title}' exported to {filepath}")
    except Exception as e:
        print(f"Failed to export report '{report.title}': {e}")
        raise


def export_report_to_json(report: NCSSReport, filepath: str) -> None:
    """
    Export NCSS report to JSON format for machine readability.
    
    Note: Plot image bytes are excluded from JSON to keep file size manageable.
    Use the Excel export for complete data including plots.
    
    Args:
        report: NCSS report to export
        filepath: Output file path
        
    Raises:
        FileNotFoundError: If directory doesn't exist and can't be created
        PermissionError: If file can't be written
    """
    try:
        # Convert report to dictionary
        report_dict = {
            'title': report.title,
            'metadata': report.metadata,
            'sections': []
        }
        
        for section in report.sections:
            section_dict = {
                'title': section.title,
                'section_type': section.section_type.value,
                'text': section.text,
                'notes': section.notes,
                'tables': [],
                'plots': []
            }
            
            # Add tables
            for table in section.tables:
                table_dict = {
                    'title': table.title,
                    'columns': table.columns,
                    'rows': table.rows,
                    'notes': table.notes
                }
                section_dict['tables'].append(table_dict)
            
            # Add plots (without image bytes for JSON)
            for plot in section.plots:
                plot_dict = {
                    'title': plot.title,
                    'description': plot.description,
                    'plot_type': plot.plot_type
                    # Note: image_bytes excluded from JSON for size
                }
                section_dict['plots'].append(plot_dict)
            
            report_dict['sections'].append(section_dict)
        
        # Write to file
        with open(filepath, 'w') as f:
            json.dump(report_dict, f, indent=2)
        
        print(f"Report '{report.title}' exported to {filepath}")
    except Exception as e:
        print(f"Failed to export report '{report.title}': {e}")
        raise


def export_tables_to_csv_batch(report: NCSSReport, output_dir: str) -> None:
    """
    Export all tables from a report to separate CSV files.
    
    Creates a directory structure with one CSV file per table,
    named using section and table titles.
    
    Args:
        report: NCSS report to export
        output_dir: Output directory path
        
    Raises:
        FileNotFoundError: If directory can't be created
        PermissionError: If files can't be written
    """
    try:
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        for section in report.sections:
            for i, table in enumerate(section.tables):
                # Create filename
                section_name = section.title.replace(' ', '_').replace('/', '_')
                table_name = table.title.replace(' ', '_').replace('/', '_')
                filename = f"{section_name}_{table_name}_{i+1}.csv"
                filepath = output_path / filename
                
                # Export table
                export_table_to_csv(table, str(filepath))
        
        print(f"All tables from report '{report.title}' exported to {output_dir}")
    except Exception as e:
        print(f"Failed to export tables: {e}")
        raise


def create_data_dictionary(report: NCSSReport) -> pd.DataFrame:
    """
    Create a data dictionary for all variables in the report.
    
    This function analyzes all tables in the report and creates
    a comprehensive data dictionary with variable information.
    
    Args:
        report: NCSS report to analyze
    
    Returns:
        pd.DataFrame: DataFrame with variable information including
                     section, table, variable name, sample values, and data type
    """
    data_dict = []
    
    for section in report.sections:
        for table in section.tables:
            for col in table.columns:
                # Get sample values from first few rows
                sample_values = []
                for row in table.rows[:5]:  # First 5 rows
                    if col in row:
                        sample_values.append(str(row[col]))
                
                # Determine data type
                is_numeric = any(
                    v.replace('.', '').replace('-', '').replace('e', '').replace('E', '').isdigit() 
                    for v in sample_values if v
                )
                
                data_dict.append({
                    'Section': section.title,
                    'Table': table.title,
                    'Variable': col,
                    'Sample_Values': '; '.join(sample_values),
                    'Data_Type': 'Numeric' if is_numeric else 'Categorical'
                })
    
    return pd.DataFrame(data_dict)


def export_data_dictionary(report: NCSSReport, filepath: str) -> None:
    """
    Export a data dictionary for the report.
    
    Args:
        report: NCSS report to analyze
        filepath: Output file path
        
    Raises:
        FileNotFoundError: If directory doesn't exist and can't be created
        PermissionError: If file can't be written
    """
    try:
        data_dict_df = create_data_dictionary(report)
        data_dict_df.to_csv(filepath, index=False)
        print(f"Data dictionary exported to {filepath}")
    except Exception as e:
        print(f"Failed to export data dictionary: {e}")
        raise


def create_regulatory_summary(report: NCSSReport) -> Dict[str, Any]:
    """
    Create a regulatory-compliant summary of the report.
    
    This function creates a structured summary suitable for
    regulatory submissions and compliance documentation.
    
    Args:
        report: NCSS report to summarize
    
    Returns:
        Dict[str, Any]: Dictionary with regulatory summary information
    """
    summary = {
        'report_title': report.title,
        'analysis_date': report.metadata.get('analysis_date', 'Not specified'),
        'software_version': report.metadata.get('software_version', 'Not specified'),
        'sections': [],
        'total_tables': 0,
        'total_plots': 0,
        'data_quality_checks': []
    }
    
    for section in report.sections:
        section_summary = {
            'title': section.title,
            'type': section.section_type.value,
            'table_count': len(section.tables),
            'plot_count': len(section.plots),
            'has_notes': section.notes is not None
        }
        summary['sections'].append(section_summary)
        summary['total_tables'] += len(section.tables)
        summary['total_plots'] += len(section.plots)
    
    # Add data quality checks
    for section in report.sections:
        if section.section_type == SectionType.DIAGNOSTICS:
            summary['data_quality_checks'].append('Diagnostic plots generated')
        if section.section_type == SectionType.ANOVA:
            summary['data_quality_checks'].append('Statistical tests performed')
    
    return summary


def export_regulatory_summary(report: NCSSReport, filepath: str) -> None:
    """
    Export a regulatory-compliant summary to JSON.
    
    Args:
        report: NCSS report to summarize
        filepath: Output file path
        
    Raises:
        FileNotFoundError: If directory doesn't exist and can't be created
        PermissionError: If file can't be written
    """
    try:
        summary = create_regulatory_summary(report)
        
        with open(filepath, 'w') as f:
            json.dump(summary, f, indent=2)
        
        print(f"Regulatory summary exported to {filepath}")
    except Exception as e:
        print(f"Failed to export regulatory summary: {e}")
        raise 